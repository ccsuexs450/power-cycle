from openpyxl import load_workbook

wb = load_workbook('test.xlsx')
ws = wb.active

ws.title = "Testing"

lines = list()
with open ("test.txt", "r") as myfile:
    for line in myfile:
        lines.append(line.strip())

x = 1
y = 1

for data in lines:
    ws.cell(row=x,column=y,value=data)
    x += 1

    
wb.save('test.xlsx')

